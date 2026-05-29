from fastapi import FastAPI, HTTPException, Depends, status
from fastapi.middleware.cors import CORSMiddleware
from fastapi.security import OAuth2PasswordBearer, OAuth2PasswordRequestForm
from pydantic import BaseModel
from typing import Optional
import sqlite3, json, os
from datetime import datetime, timedelta
from jose import JWTError, jwt
import bcrypt as _bcrypt

app = FastAPI(title="Agon API")

ALLOWED_ORIGINS = os.environ.get(
    "ALLOWED_ORIGINS",
    "https://agon.savo.us,http://localhost,http://localhost:3800"
).split(",")

app.add_middleware(
    CORSMiddleware,
    allow_origins=ALLOWED_ORIGINS,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

DB_PATH           = os.environ.get("DB_PATH", "/data/sessions.db")
SECRET_KEY        = os.environ.get("SECRET_KEY", "agon-change-this-in-production")
ALGORITHM         = "HS256"
TOKEN_EXPIRE_DAYS = 30

oauth2 = OAuth2PasswordBearer(tokenUrl="/auth/login")

ADMIN_USER = os.environ.get("ADMIN_USER", "manny")
ADMIN_PASS = os.environ.get("ADMIN_PASS", "pR0m3th3us")
DEMO_USER  = "demo"
DEMO_PASS  = "demo"


# ── DB ────────────────────────────────────────────────────────
def get_db():
    os.makedirs(os.path.dirname(DB_PATH), exist_ok=True)
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn


def migrate_sessions_to_per_set_time(conn):
    """
    One-time migration: move ex.time down to each set as set.time.
    After migration, ex.time is removed (density recalculated from sets).
    Idempotent — skips sets that already have a time field.
    """
    rows = conn.execute("SELECT id, exercises FROM sessions").fetchall()
    updated = 0
    for r in rows:
        exercises = json.loads(r["exercises"])
        changed = False
        for ex in exercises:
            ex_time = ex.get("time", 10)
            new_sets = []
            for s in ex.get("sets", []):
                if "time" not in s:
                    s["time"] = ex_time
                    changed = True
                new_sets.append(s)
            ex["sets"] = new_sets
            # Recalculate density from per-set data
            if changed:
                total_density = sum(
                    (s.get("vol", 0) / s.get("time", 1))
                    for s in new_sets if s.get("time", 0) > 0
                )
                ex["density"] = round(total_density, 2)
                ex["totalVol"] = round(sum(s.get("vol", 0) for s in new_sets), 1)
        if changed:
            # Recalculate session RD
            bw = conn.execute("SELECT bw FROM sessions WHERE id=?", (r["id"],)).fetchone()["bw"]
            total_vol  = sum(s.get("vol", 0)  for ex in exercises for s in ex.get("sets", []))
            total_time = sum(s.get("time", 0) for ex in exercises for s in ex.get("sets", []) if s.get("time", 0) > 0)
            total_density = round(total_vol / total_time, 2) if total_time > 0 else 0
            rd = round(total_vol / (total_time * bw), 2) if (total_time > 0 and bw) else 0
            conn.execute(
                "UPDATE sessions SET exercises=?, rd=?, total_density=? WHERE id=?",
                (json.dumps(exercises), rd, total_density, r["id"])
            )
            updated += 1
    if updated:
        conn.commit()
        print(f"Migrated {updated} sessions to per-set time model")


def init_db():
    conn = get_db()

    conn.execute("""
        CREATE TABLE IF NOT EXISTS users (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            username   TEXT NOT NULL UNIQUE,
            hashed_pw  TEXT NOT NULL,
            role       TEXT NOT NULL DEFAULT 'guest',
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)

    conn.execute("""
        CREATE TABLE IF NOT EXISTS sessions (
            id            INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id       INTEGER NOT NULL DEFAULT 1,
            date          TEXT NOT NULL,
            bw            REAL NOT NULL,
            rd            REAL NOT NULL,
            total_density REAL NOT NULL,
            exercises     TEXT NOT NULL,
            created_at    TEXT DEFAULT (datetime('now')),
            UNIQUE(user_id, date)
        )
    """)

    cols = [r[1] for r in conn.execute("PRAGMA table_info(sessions)").fetchall()]
    if "user_id" not in cols:
        conn.execute("ALTER TABLE sessions ADD COLUMN user_id INTEGER NOT NULL DEFAULT 1")
    if "notes" not in cols:
        conn.execute("ALTER TABLE sessions ADD COLUMN notes TEXT NOT NULL DEFAULT ''")

    conn.commit()

    # Seed admin
    if not conn.execute("SELECT id FROM users WHERE username=?", (ADMIN_USER,)).fetchone():
        conn.execute(
            "INSERT INTO users (username, hashed_pw, role) VALUES (?,?,?)",
            (ADMIN_USER, _bcrypt.hashpw(ADMIN_PASS.encode(), _bcrypt.gensalt()).decode(), "admin")
        )
    # Seed demo
    if not conn.execute("SELECT id FROM users WHERE username=?", (DEMO_USER,)).fetchone():
        conn.execute(
            "INSERT INTO users (username, hashed_pw, role) VALUES (?,?,?)",
            (DEMO_USER, _bcrypt.hashpw(DEMO_PASS.encode(), _bcrypt.gensalt()).decode(), "demo")
        )

    conn.execute("""
        CREATE TABLE IF NOT EXISTS exercises (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            name       TEXT NOT NULL UNIQUE,
            alias      TEXT,
            tool       TEXT NOT NULL DEFAULT 'Bar',
            mult       REAL NOT NULL DEFAULT 2.0,
            muscles    TEXT NOT NULL DEFAULT '[]',
            day        TEXT,
            load_hint  TEXT,
            is_bw      INTEGER NOT NULL DEFAULT 0,
            sort_order INTEGER NOT NULL DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now'))
        )
    """)

    # Per-user AI settings, profile info, and preferences
    conn.execute("""
        CREATE TABLE IF NOT EXISTS user_settings (
            user_id      INTEGER PRIMARY KEY,
            ai_key_enc   TEXT,
            ai_model     TEXT DEFAULT 'gemini-2.5-flash',
            first_name   TEXT,
            last_name    TEXT,
            dob          TEXT,
            gender       TEXT,
            week_start   TEXT DEFAULT 'Saturday',
            height_in    REAL,
            target_bw    REAL,
            updated_at   TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    # Migrate existing rows to add new profile columns if they don't exist
    for col, default in [('first_name','NULL'),('last_name','NULL'),('dob','NULL'),
                         ('gender','NULL'),('week_start',"'Saturday'"),
                         ('height_in','NULL'),('target_bw','NULL'),
                         ('activity_level',"'1.55'"),
                         ('onboarded',"'0'")]:
        try:
            conn.execute(f"ALTER TABLE user_settings ADD COLUMN {col} TEXT DEFAULT {default}")
            conn.commit()
        except Exception:
            pass

    # Supplement/peptide protocols table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS protocols (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            name       TEXT NOT NULL,
            dose       TEXT,
            frequency  TEXT,
            notes      TEXT,
            sort_order INTEGER DEFAULT 0,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)
    conn.commit()

    # Training phases table
    conn.execute("""
        CREATE TABLE IF NOT EXISTS phases (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            phase_type TEXT NOT NULL,
            start_date TEXT NOT NULL,
            end_date   TEXT,
            notes      TEXT,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)

    # Persisted Insights conversation history (one row per message)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS insights_messages (
            id         INTEGER PRIMARY KEY AUTOINCREMENT,
            user_id    INTEGER NOT NULL,
            role       TEXT NOT NULL,
            content    TEXT NOT NULL,
            created_at TEXT DEFAULT (datetime('now')),
            FOREIGN KEY (user_id) REFERENCES users(id)
        )
    """)

    conn.commit()

    # Run per-set time migration (only once, gated by a flag in meta table)
    conn.execute("""
        CREATE TABLE IF NOT EXISTS meta (
            key   TEXT PRIMARY KEY,
            value TEXT NOT NULL
        )
    """)
    conn.commit()
    flag = conn.execute("SELECT value FROM meta WHERE key='per_set_time_migration_done'").fetchone()
    if not flag:
        migrate_sessions_to_per_set_time(conn)
        conn.execute("INSERT OR REPLACE INTO meta (key, value) VALUES (?, ?)",
                     ('per_set_time_migration_done', '1'))
        conn.commit()
    conn.close()


init_db()


# ── Auth ──────────────────────────────────────────────────────
def verify_password(plain: str, hashed: str) -> bool:
    return _bcrypt.checkpw(plain.encode(), hashed.encode())

def hash_password(plain: str) -> str:
    return _bcrypt.hashpw(plain.encode(), _bcrypt.gensalt()).decode()


# ── AI key encryption (Fernet, derived from SECRET_KEY) ───────
import base64 as _b64, hashlib as _hashlib
from cryptography.fernet import Fernet, InvalidToken

def _fernet():
    # Derive a stable 32-byte urlsafe key from SECRET_KEY
    digest = _hashlib.sha256(SECRET_KEY.encode()).digest()
    return Fernet(_b64.urlsafe_b64encode(digest))

def encrypt_secret(plain: str) -> str:
    return _fernet().encrypt(plain.encode()).decode()

def decrypt_secret(token: str) -> str:
    try:
        return _fernet().decrypt(token.encode()).decode()
    except (InvalidToken, Exception):
        return ""

def create_token(data: dict):
    payload = data.copy()
    payload["exp"] = datetime.utcnow() + timedelta(days=TOKEN_EXPIRE_DAYS)
    return jwt.encode(payload, SECRET_KEY, algorithm=ALGORITHM)

def get_user(conn, username: str):
    return conn.execute("SELECT * FROM users WHERE username=?", (username,)).fetchone()

async def current_user(token: str = Depends(oauth2)):
    err = HTTPException(status_code=401, detail="Invalid or expired token",
                        headers={"WWW-Authenticate": "Bearer"})
    try:
        payload = jwt.decode(token, SECRET_KEY, algorithms=[ALGORITHM])
        username = payload.get("sub")
        if not username:
            raise err
    except JWTError:
        raise err
    conn = get_db()
    user = get_user(conn, username)
    conn.close()
    if not user:
        raise err
    return dict(user)

async def admin_only(user=Depends(current_user)):
    if user["role"] != "admin":
        raise HTTPException(status_code=403, detail="Admin access required")
    return user

def data_user_id(user: dict) -> int:
    if user["role"] == "demo":
        conn = get_db()
        admin = conn.execute("SELECT id FROM users WHERE role='admin' LIMIT 1").fetchone()
        conn.close()
        return admin["id"] if admin else 1
    return user["id"]


# ── Models ────────────────────────────────────────────────────
class SetData(BaseModel):
    reps: float
    rawLoad: float
    trueLbs: float
    vol: float
    time: float = 2.0          # per-set time in minutes

class ExerciseData(BaseModel):
    name: str
    sets: list[SetData]
    totalVol: float
    density: float
    tool: Optional[str] = "Bar"
    mult: Optional[float] = 2.0
    isBW: Optional[bool] = False
    time: Optional[float] = None  # kept for legacy compat, ignored in calc

class ExerciseIn(BaseModel):
    name: str
    alias: Optional[str] = ''
    tool: str = 'Bar'
    mult: float = 2.0
    muscles: list[str] = []
    day: Optional[str] = ''
    load_hint: Optional[str] = ''
    is_bw: bool = False
    sort_order: int = 0

class SessionIn(BaseModel):
    date: str
    bw: float
    rd: float
    total_density: float
    exercises: list[ExerciseData]
    notes: Optional[str] = ''

class UserCreate(BaseModel):
    username: str
    password: str
    role: str = "guest"


class ProfileIn(BaseModel):
    first_name: Optional[str] = None
    last_name:  Optional[str] = None
    dob:        Optional[str] = None        # YYYY-MM-DD
    gender:     Optional[str] = None
    week_start: Optional[str] = None        # Monday-Sunday
    height_in:      Optional[float] = None      # inches
    target_bw:      Optional[float] = None      # lbs
    activity_level: Optional[str]   = None      # TDEE multiplier string
    onboarded:      Optional[str]   = None      # "1" once user completes onboarding

class ProtocolIn(BaseModel):
    name:      str
    dose:      Optional[str] = None
    frequency: Optional[str] = None
    notes:     Optional[str] = None

class ProtocolsBulkIn(BaseModel):
    protocols: list

class PhaseIn(BaseModel):
    phase_type: str
    start_date: str
    end_date:   Optional[str] = None
    notes:      Optional[str] = None

class AISettingsIn(BaseModel):
    ai_key: Optional[str] = None      # plaintext key from user; None = don't change
    ai_model: Optional[str] = None

class InsightsQuery(BaseModel):
    question: str


# ── Auth endpoints ────────────────────────────────────────────
@app.get("/health")
def health():
    return {"status": "ok"}

@app.post("/auth/login")
def login(form: OAuth2PasswordRequestForm = Depends()):
    conn = get_db()
    user = get_user(conn, form.username)
    conn.close()
    if not user or not verify_password(form.password, user["hashed_pw"]):
        raise HTTPException(status_code=401, detail="Incorrect username or password")
    token = create_token({"sub": user["username"], "role": user["role"]})
    return {"access_token": token, "token_type": "bearer",
            "role": user["role"], "username": user["username"]}

@app.get("/auth/me")
def me(user=Depends(current_user)):
    return {"username": user["username"], "role": user["role"], "id": user["id"]}


# ── Admin ─────────────────────────────────────────────────────
@app.get("/admin/users")
def list_users(user=Depends(admin_only)):
    conn = get_db()
    rows = conn.execute("SELECT id, username, role, created_at FROM users ORDER BY id").fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/admin/users")
def create_user(body: UserCreate, user=Depends(admin_only)):
    conn = get_db()
    if conn.execute("SELECT id FROM users WHERE username=?", (body.username,)).fetchone():
        conn.close()
        raise HTTPException(status_code=400, detail="Username already exists")
    conn.execute("INSERT INTO users (username, hashed_pw, role) VALUES (?,?,?)",
                 (body.username, hash_password(body.password), body.role))
    conn.commit()
    conn.close()
    return {"status": "created", "username": body.username}

@app.delete("/admin/users/{username}")
def delete_user(username: str, user=Depends(admin_only)):
    if username == ADMIN_USER:
        raise HTTPException(status_code=400, detail="Cannot delete system accounts")
    conn = get_db()
    conn.execute("DELETE FROM users WHERE username=?", (username,))
    conn.commit()
    conn.close()
    return {"status": "deleted", "username": username}

@app.put("/admin/users/{username}/password")
def change_password(username: str, body: dict, user=Depends(admin_only)):
    new_pw = body.get("password", "")
    if len(new_pw) < 4:
        raise HTTPException(status_code=400, detail="Password too short")
    conn = get_db()
    conn.execute("UPDATE users SET hashed_pw=? WHERE username=?",
                 (hash_password(new_pw), username))
    conn.commit()
    conn.close()
    return {"status": "updated"}


# ── Sessions ──────────────────────────────────────────────────
@app.get("/sessions")
def get_sessions(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT * FROM sessions WHERE user_id=? ORDER BY date ASC", (uid,)
    ).fetchall()
    conn.close()
    return [{"id": r["id"], "date": r["date"], "bw": r["bw"], "rd": r["rd"],
             "total_density": r["total_density"], "notes": r["notes"] or "",
             "exercises": json.loads(r["exercises"])} for r in rows]

@app.post("/sessions")
def save_session(session: SessionIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo account is read-only")
    uid = user["id"]

    # Recalculate RD server-side using correct formula:
    # RD = total_session_vol / (total_session_time * bodyweight)
    exercises_out = []
    total_vol  = 0.0
    total_time = 0.0
    for ex in session.exercises:
        ex_vol  = sum(s.vol  for s in ex.sets)
        ex_time = sum(s.time for s in ex.sets if s.time > 0)
        total_vol  += ex_vol
        total_time += ex_time
        exercises_out.append({
            **ex.dict(),
            "totalVol": round(ex_vol, 1),
            "density":  round(ex_vol / ex_time, 2) if ex_time > 0 else 0,
        })

    total_density = round(total_vol / total_time, 2) if total_time > 0 else 0
    rd = round(total_vol / (total_time * session.bw), 2) if (total_time > 0 and session.bw) else 0

    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO sessions (user_id, date, bw, rd, total_density, exercises, notes)
            VALUES (?,?,?,?,?,?,?)
            ON CONFLICT(user_id, date) DO UPDATE SET
                bw=excluded.bw, rd=excluded.rd,
                total_density=excluded.total_density, exercises=excluded.exercises,
                notes=excluded.notes
        """, (uid, session.date, session.bw, rd, round(total_density, 2),
              json.dumps(exercises_out), session.notes or ""))
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=500, detail=str(e))
    conn.close()
    return {"status": "saved", "date": session.date, "rd": rd}

@app.delete("/sessions/{date}")
def delete_session(date: str, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo account is read-only")
    conn = get_db()
    conn.execute("DELETE FROM sessions WHERE user_id=? AND date=?", (user["id"], date))
    conn.commit()
    conn.close()
    return {"status": "deleted", "date": date}


# ── Trend ─────────────────────────────────────────────────────
@app.get("/trend")
def get_trend(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd FROM sessions WHERE user_id=? ORDER BY date ASC", (uid,)
    ).fetchall()
    conn.close()

    # Determine week_start preference (defaults to Saturday)
    conn2 = get_db()
    settings_row = conn2.execute("SELECT week_start FROM user_settings WHERE user_id=?", (uid,)).fetchone()
    conn2.close()
    week_start_pref = (settings_row["week_start"] if settings_row and settings_row["week_start"] else "Saturday")

    # weekday(): Mon=0, Tue=1, Wed=2, Thu=3, Fri=4, Sat=5, Sun=6
    # Week ends on the day BEFORE week_start
    DAY_MAP = {"Monday":0,"Tuesday":1,"Wednesday":2,"Thursday":3,"Friday":4,"Saturday":5,"Sunday":6}
    start_wd = DAY_MAP.get(week_start_pref, 5)  # default Saturday=5
    end_wd = (start_wd - 1) % 7  # week-end day

    by_week: dict = {}
    for r in rows:
        try:
            d = datetime.strptime(r["date"], "%Y-%m-%d")
        except ValueError:
            continue
        delta = (end_wd - d.weekday()) % 7
        week_end = d + timedelta(days=delta)
        key = f"{week_end.month}/{week_end.day}/{week_end.year}"
        if key not in by_week:
            by_week[key] = []
        by_week[key].append({"rd": r["rd"], "wt": r["bw"]})

    result = []
    for week, entries in by_week.items():
        result.append({
            "week": week,
            "rd":   round(sum(e["rd"] for e in entries) / len(entries), 2),
            "wt":   round(sum(e["wt"] for e in entries) / len(entries), 1),
        })
    result.sort(key=lambda w: datetime.strptime(w["week"], "%m/%d/%Y"))
    return result


# ── Progress ──────────────────────────────────────────────────
@app.get("/progress/{exercise_name}")
def get_progress(exercise_name: str, user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd, exercises FROM sessions WHERE user_id=? ORDER BY date ASC", (uid,)
    ).fetchall()
    conn.close()

    result = []
    for r in rows:
        exercises = json.loads(r["exercises"])
        match = next((e for e in exercises
                      if e.get("name","").lower() == exercise_name.lower()), None)
        if not match:
            continue
        sets = match.get("sets", [])
        if not sets:
            continue
        top_set  = max(sets, key=lambda s: s.get("trueLbs", 0))
        total_vol = sum(s.get("vol", 0) for s in sets)
        result.append({
            "date":     r["date"],
            "bw":       r["bw"],
            "rd":       r["rd"],
            "topLoad":  round(top_set.get("trueLbs", 0), 1),
            "topReps":  int(top_set.get("reps", 0)),
            "totalVol": round(total_vol, 1),
            "density":  round(match.get("density", 0), 2),
            "sets":     len(sets),
        })
    return result


@app.get("/exercises/bank")
def get_exercise_bank(user=Depends(current_user)):
    conn = get_db()
    rows = conn.execute("SELECT * FROM exercises ORDER BY sort_order, name").fetchall()
    conn.close()
    return [{"id":r["id"],"name":r["name"],"alias":r["alias"],"tool":r["tool"],
             "mult":r["mult"],"muscles":json.loads(r["muscles"]),"day":r["day"],
             "loadHint":r["load_hint"],"isBW":bool(r["is_bw"]),"sortOrder":r["sort_order"]} for r in rows]

@app.post("/exercises/bank")
def add_exercise(body: ExerciseIn, user=Depends(current_user)):
    if user["role"] not in ("admin", "guest"):
        raise HTTPException(status_code=403, detail="Read-only account")
    conn = get_db()
    try:
        conn.execute("""
            INSERT INTO exercises (name, alias, tool, mult, muscles, day, load_hint, is_bw, sort_order)
            VALUES (?,?,?,?,?,?,?,?,?)
            ON CONFLICT(name) DO UPDATE SET
                alias=excluded.alias, tool=excluded.tool, mult=excluded.mult,
                muscles=excluded.muscles, day=excluded.day, load_hint=excluded.load_hint,
                is_bw=excluded.is_bw, sort_order=excluded.sort_order
        """, (body.name, body.alias, body.tool, body.mult, json.dumps(body.muscles),
              body.day, body.load_hint, int(body.is_bw), body.sort_order))
        conn.commit()
    except Exception as e:
        conn.close()
        raise HTTPException(status_code=400, detail=str(e))
    conn.close()
    return {"status": "upserted", "name": body.name}

@app.put("/exercises/bank/{ex_name}")
def update_exercise(ex_name: str, body: ExerciseIn, user=Depends(current_user)):
    if user["role"] not in ("admin", "guest"):
        raise HTTPException(status_code=403, detail="Read-only account")
    conn = get_db()
    result = conn.execute(
        "UPDATE exercises SET alias=?, tool=?, mult=?, muscles=?, day=?, load_hint=?, is_bw=?, sort_order=? WHERE name=?",
        (body.alias, body.tool, body.mult, json.dumps(body.muscles),
         body.day, body.load_hint, int(body.is_bw), body.sort_order, ex_name))
    if result.rowcount == 0:
        # Doesn't exist — insert it
        conn.execute(
            "INSERT INTO exercises (name, alias, tool, mult, muscles, day, load_hint, is_bw, sort_order) VALUES (?,?,?,?,?,?,?,?,?)",
            (ex_name, body.alias, body.tool, body.mult, json.dumps(body.muscles),
             body.day, body.load_hint, int(body.is_bw), body.sort_order))
    conn.commit()
    conn.close()
    return {"status": "upserted"}

@app.delete("/exercises/bank/{ex_name}")
def delete_exercise(ex_name: str, user=Depends(current_user)):
    if user["role"] not in ("admin", "guest"):
        raise HTTPException(status_code=403, detail="Read-only account")
    conn = get_db()
    conn.execute("DELETE FROM exercises WHERE name=?", (ex_name,))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

@app.get("/exercises/logged")
def get_logged_exercises(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT exercises FROM sessions WHERE user_id=?", (uid,)
    ).fetchall()
    conn.close()
    names = set()
    for r in rows:
        for ex in json.loads(r["exercises"]):
            if ex.get("name"):
                names.add(ex["name"])
    return sorted(names)

# ── Profile endpoints ────────────────────────────────────────
@app.get("/profile")
def get_profile(user=Depends(current_user)):
    conn = get_db()
    try:
        row = conn.execute(
            "SELECT first_name, last_name, dob, gender, week_start, height_in, target_bw, activity_level, onboarded FROM user_settings WHERE user_id=?",
            (user["id"],)
        ).fetchone()
    except Exception:
        # Columns may not exist yet — fall back to partial query
        try:
            row = conn.execute(
                "SELECT first_name, last_name, dob, gender, week_start FROM user_settings WHERE user_id=?",
                (user["id"],)
            ).fetchone()
        except Exception:
            row = None
    conn.close()
    if not row:
        return {"first_name":"","last_name":"","dob":"","gender":"","week_start":"Saturday","height_in":None,"target_bw":None}
    return {
        "first_name": row["first_name"] or "",
        "last_name":  row["last_name"]  or "",
        "dob":        row["dob"]        or "",
        "gender":     row["gender"]     or "",
        "week_start":       row["week_start"] or "Saturday",
        "height_in":        row["height_in"]        if "height_in"        in row.keys() else None,
        "target_bw":        row["target_bw"]        if "target_bw"        in row.keys() else None,
        "activity_level":   row["activity_level"]   if "activity_level"   in row.keys() else "1.55",
        "onboarded":        row["onboarded"]        if "onboarded"        in row.keys() else "0",
    }

@app.post("/profile")
def save_profile(body: ProfileIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot save profile data")
    conn = get_db()
    existing = conn.execute("SELECT user_id FROM user_settings WHERE user_id=?", (user["id"],)).fetchone()
    if existing:
        conn.execute("""
            UPDATE user_settings SET
                first_name=COALESCE(?,first_name),
                last_name=COALESCE(?,last_name),
                dob=COALESCE(?,dob),
                gender=COALESCE(?,gender),
                week_start=COALESCE(?,week_start),
                height_in=COALESCE(?,height_in),
                target_bw=COALESCE(?,target_bw),
                activity_level=COALESCE(?,activity_level),
                onboarded=COALESCE(?,onboarded),
                updated_at=datetime('now')
            WHERE user_id=?
        """, (body.first_name, body.last_name, body.dob, body.gender, body.week_start,
                body.height_in, body.target_bw, body.activity_level, body.onboarded, user["id"]))
    else:
        conn.execute("""
            INSERT INTO user_settings (user_id, first_name, last_name, dob, gender, week_start, height_in, target_bw)
            VALUES (?,?,?,?,?,?,?,?)
        """, (user["id"], body.first_name, body.last_name, body.dob, body.gender,
                body.week_start or "Saturday", body.height_in, body.target_bw, body.activity_level or "1.55"))
    conn.commit()
    conn.close()
    return {"status": "saved"}


# ── AI Settings endpoints ─────────────────────────────────────
@app.get("/ai/settings")
def get_ai_settings(user=Depends(current_user)):
    """Return whether a key is set + masked preview + model. Never returns the raw key."""
    conn = get_db()
    row = conn.execute("SELECT ai_key_enc, ai_model FROM user_settings WHERE user_id=?",
                       (user["id"],)).fetchone()
    conn.close()
    if not row or not row["ai_key_enc"]:
        return {"has_key": False, "masked": "", "model": (row["ai_model"] if row else "gemini-1.5-flash")}
    plain = decrypt_secret(row["ai_key_enc"])
    masked = ("•" * max(0, len(plain) - 4) + plain[-4:]) if plain else ""
    return {"has_key": bool(plain), "masked": masked, "model": row["ai_model"] or "gemini-1.5-flash"}

@app.post("/ai/settings")
def save_ai_settings(body: AISettingsIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot store API keys")
    conn = get_db()
    existing = conn.execute("SELECT user_id FROM user_settings WHERE user_id=?",
                            (user["id"],)).fetchone()
    # Build the update — always write both fields atomically
    key_enc = None
    if body.ai_key is not None and body.ai_key.strip():
        key_enc = encrypt_secret(body.ai_key.strip())
    if existing:
        cur = conn.execute("SELECT ai_key_enc, ai_model FROM user_settings WHERE user_id=?",
                           (user["id"],)).fetchone()
        final_key   = key_enc if key_enc is not None else cur["ai_key_enc"]
        final_model = body.ai_model if body.ai_model else (cur["ai_model"] or "gemini-2.5-flash")
        conn.execute(
            "UPDATE user_settings SET ai_key_enc=?, ai_model=?, updated_at=datetime('now') WHERE user_id=?",
            (final_key, final_model, user["id"])
        )
    else:
        conn.execute("INSERT INTO user_settings (user_id, ai_key_enc, ai_model) VALUES (?,?,?)",
                     (user["id"], key_enc, body.ai_model or "gemini-2.5-flash"))
    conn.commit()
    conn.close()
    return {"status": "saved"}

@app.delete("/ai/settings")
def delete_ai_settings(user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts have no stored key")
    conn = get_db()
    conn.execute("UPDATE user_settings SET ai_key_enc=NULL, updated_at=datetime('now') WHERE user_id=?",
                 (user["id"],))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

# ── Insights: build a compact data context for the LLM ────────
def build_training_context(uid: int) -> str:
    """Produce a compact text summary of the user's training data for the LLM."""
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd, total_density, exercises, notes FROM sessions WHERE user_id=? ORDER BY date",
        (uid,)
    ).fetchall()
    conn.close()
    if not rows:
        return "No training sessions logged yet."

    lines = []
    # Load profile for richer AI context
    conn_p = get_db()
    try:
        prof = conn_p.execute(
            "SELECT first_name, last_name, dob, gender, week_start, height_in, target_bw FROM user_settings WHERE user_id=?",
            (uid,)
        ).fetchone()
    except Exception:
        prof = conn_p.execute(
            "SELECT first_name, last_name, dob, gender, week_start FROM user_settings WHERE user_id=?",
            (uid,)
        ).fetchone()
    conn_p.close()

    def prof_get(key, default=None):
        try:
            return prof[key]
        except (IndexError, TypeError):
            return default

    if prof:
        name_str = " ".join(filter(None, [prof_get("first_name"), prof_get("last_name")])) or "User"
        age_str = ""
        if prof_get("dob"):
            try:
                from datetime import date as _date
                dob = datetime.strptime(prof_get("dob"), "%Y-%m-%d").date()
                age_str = f", age {(_date.today().year - dob.year - ((_date.today().month, _date.today().day) < (dob.month, dob.day)))}"
            except Exception:
                pass
        lines.append(f"Athlete: {name_str}{age_str}, gender: {prof_get('gender') or 'not specified'}")
        lines.append(f"Training week: {prof_get('week_start') or 'Saturday'} through {['Sunday','Monday','Tuesday','Wednesday','Thursday','Friday','Saturday'][(['Monday','Tuesday','Wednesday','Thursday','Friday','Saturday','Sunday'].index(prof_get('week_start') or 'Saturday') + 6) % 7]}")
        if prof_get("height_in"):
            h = float(prof_get("height_in"))
            lines.append(f"Height: {int(h//12)}'{int(h%12)}\" ({h} in)")
        if prof_get("target_bw"):
            lines.append(f"Target bodyweight: {prof_get('target_bw')} lbs")
        lines.append("")

    # Add active phase
    conn_ph = get_db()
    active_phase = conn_ph.execute(
        "SELECT phase_type, start_date FROM phases WHERE user_id=? AND end_date IS NULL ORDER BY start_date DESC LIMIT 1",
        (uid,)
    ).fetchone()
    conn_ph.close()
    if active_phase:
        lines.append(f"Current training phase: {active_phase['phase_type']} (since {active_phase['start_date']})")
        lines.append("")

    lines.append(f"Total sessions: {len(rows)}")
    lines.append(f"Date range: {rows[0]['date']} to {rows[-1]['date']}")
    lines.append("")
    lines.append("RD = Relative Density = total session volume / (total time x bodyweight).")
    lines.append("Higher RD means more work done per unit time per pound of bodyweight.")
    lines.append("")
    lines.append("SESSION LOG (date | bodyweight | RD | exercises[load x reps]):")

    for r in rows:
        try:
            exs = json.loads(r["exercises"])
        except Exception:
            exs = []
        ex_parts = []
        for ex in exs:
            sets = ex.get("sets", [])
            if not sets:
                continue
            # Compact: name maxload xtotalreps
            max_load = max((s.get("trueLbs") or s.get("rawLoad") or 0) for s in sets)
            total_reps = sum(s.get("reps", 0) for s in sets)
            ex_parts.append(f"{ex.get('name','?')} {max_load:g}x{total_reps:g}")
        note = f" | note: {r['notes']}" if r["notes"] else ""
        lines.append(f"{r['date']} | {r['bw']:g}lb | RD {r['rd']:.2f} | {'; '.join(ex_parts)}{note}")

    return "\n".join(lines)


SYSTEM_PROMPT = """You are an analytical strength-training assistant inside an app called Agon.
The user tracks workouts using a metric called RD (Relative Density). You have access to their
full session history below. Answer their questions directly and concisely, grounding every claim
in the actual data. Use specific dates, loads, and numbers. If the data doesn't support an answer,
say so plainly. Do not invent data. Keep answers tight and useful — this is a knowledgeable user
who wants signal, not filler. When discussing trends, cite the specific sessions that show them."""


def _is_claude_model(model: str) -> bool:
    return model.startswith("claude-")

async def call_llm(api_key: str, model: str, context: str, history: list, question: str) -> str:
    """Route to Gemini or Anthropic depending on the model name."""
    import httpx
    if _is_claude_model(model):
        return await _call_anthropic(api_key, model, context, history, question)
    return await _call_gemini(api_key, model, context, history, question)

async def _call_gemini(api_key: str, model: str, context: str, history: list, question: str) -> str:
    """Call the Gemini API with training context + conversation history."""
    import httpx
    url = f"https://generativelanguage.googleapis.com/v1/models/{model}:generateContent"
    contents = []
    contents.append({
        "role": "user",
        "parts": [{"text": f"{SYSTEM_PROMPT}\n\n=== TRAINING DATA ===\n{context}\n\n=== END DATA ===\n\nAcknowledge you have the data and are ready."}]
    })
    contents.append({
        "role": "model",
        "parts": [{"text": "I have your full training history loaded and I'm ready to analyze it. What would you like to know?"}]
    })
    for msg in history:
        contents.append({
            "role": "user" if msg["role"] == "user" else "model",
            "parts": [{"text": msg["content"]}]
        })
    contents.append({"role": "user", "parts": [{"text": question}]})
    payload = {"contents": contents}
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(url, params={"key": api_key}, json=payload)
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Gemini API error ({resp.status_code}): {resp.text[:300]}")
        data = resp.json()
        try:
            return data["candidates"][0]["content"]["parts"][0]["text"]
        except (KeyError, IndexError):
            raise HTTPException(status_code=502, detail="Gemini returned an unexpected response shape")

async def _call_anthropic(api_key: str, model: str, context: str, history: list, question: str) -> str:
    """Call the Anthropic Messages API with training context + conversation history."""
    import httpx
    url = "https://api.anthropic.com/v1/messages"
    messages = []
    # Prior conversation turns
    for msg in history:
        messages.append({
            "role": "user" if msg["role"] == "user" else "assistant",
            "content": msg["content"]
        })
    messages.append({"role": "user", "content": question})
    payload = {
        "model": model,
        "max_tokens": 1024,
        "system": f"{SYSTEM_PROMPT}\n\n=== TRAINING DATA ===\n{context}\n\n=== END DATA ===",
        "messages": messages
    }
    headers = {
        "x-api-key": api_key,
        "anthropic-version": "2023-06-01",
        "content-type": "application/json"
    }
    async with httpx.AsyncClient(timeout=60) as client:
        resp = await client.post(url, headers=headers, json=payload)
        if resp.status_code != 200:
            raise HTTPException(status_code=502, detail=f"Anthropic API error ({resp.status_code}): {resp.text[:300]}")
        data = resp.json()
        try:
            return data["content"][0]["text"]
        except (KeyError, IndexError):
            raise HTTPException(status_code=502, detail="Anthropic returned an unexpected response shape")


@app.get("/insights/history")
def get_insights_history(user=Depends(current_user)):
    """Return persisted conversation. Demo gets nothing (shared identity)."""
    if user["role"] == "demo":
        return {"messages": []}
    conn = get_db()
    rows = conn.execute(
        "SELECT role, content, created_at FROM insights_messages WHERE user_id=? ORDER BY id",
        (user["id"],)
    ).fetchall()
    conn.close()
    return {"messages": [dict(r) for r in rows]}

@app.delete("/insights/history")
def clear_insights_history(user=Depends(current_user)):
    if user["role"] == "demo":
        return {"status": "noop"}
    conn = get_db()
    conn.execute("DELETE FROM insights_messages WHERE user_id=?", (user["id"],))
    conn.commit()
    conn.close()
    return {"status": "cleared"}

@app.post("/insights/ask")
async def insights_ask(body: InsightsQuery, user=Depends(current_user)):
    q = (body.question or "").strip()
    if not q:
        raise HTTPException(status_code=400, detail="Empty question")

    # Get the user's key + model
    conn = get_db()
    settings = conn.execute("SELECT ai_key_enc, ai_model FROM user_settings WHERE user_id=?",
                            (user["id"],)).fetchone()
    conn.close()
    if not settings or not settings["ai_key_enc"]:
        raise HTTPException(status_code=400, detail="No API key configured. Add one in Settings.")
    api_key = decrypt_secret(settings["ai_key_enc"])
    if not api_key:
        raise HTTPException(status_code=400, detail="Stored key could not be decrypted. Re-enter it in Settings.")
    model = settings["ai_model"] or "gemini-2.5-flash"

    # Build context from the data the user is allowed to see
    uid = data_user_id(user)
    context = build_training_context(uid)

    # Load prior conversation (demo has none)
    history = []
    if user["role"] != "demo":
        conn = get_db()
        rows = conn.execute(
            "SELECT role, content FROM insights_messages WHERE user_id=? ORDER BY id",
            (user["id"],)
        ).fetchall()
        conn.close()
        history = [dict(r) for r in rows]

    answer = await call_llm(api_key, model, context, history, q)

    # Persist both turns (not for demo)
    if user["role"] != "demo":
        conn = get_db()
        conn.execute("INSERT INTO insights_messages (user_id, role, content) VALUES (?,?,?)",
                     (user["id"], "user", q))
        conn.execute("INSERT INTO insights_messages (user_id, role, content) VALUES (?,?,?)",
                     (user["id"], "model", answer))
        conn.commit()
        conn.close()

    return {"answer": answer}

@app.post("/ai/test")
async def test_ai_key(body: AISettingsIn, user=Depends(current_user)):
    """Test a Gemini key with a trivial call. Uses provided key, or stored key if none given."""
    import httpx
    key = (body.ai_key or "").strip()
    model = body.ai_model or "gemini-1.5-flash"
    if not key:
        conn = get_db()
        row = conn.execute("SELECT ai_key_enc FROM user_settings WHERE user_id=?",
                           (user["id"],)).fetchone()
        conn.close()
        if row and row["ai_key_enc"]:
            key = decrypt_secret(row["ai_key_enc"])
    if not key:
        raise HTTPException(status_code=400, detail="No key to test")

    try:
        import httpx
        if _is_claude_model(model):
            url = "https://api.anthropic.com/v1/messages"
            headers = {"x-api-key": key, "anthropic-version": "2023-06-01", "content-type": "application/json"}
            payload = {"model": model, "max_tokens": 10, "messages": [{"role": "user", "content": "Reply OK"}]}
            async with httpx.AsyncClient(timeout=20) as client:
                resp = await client.post(url, headers=headers, json=payload)
        else:
            url = f"https://generativelanguage.googleapis.com/v1/models/{model}:generateContent"
            payload = {"contents": [{"role": "user", "parts": [{"text": "Reply with just: OK"}]}]}
            async with httpx.AsyncClient(timeout=20) as client:
                resp = await client.post(url, params={"key": key}, json=payload)
    except Exception as e:
        raise HTTPException(status_code=502, detail=f"Connection failed: {str(e)[:200]}")
    if resp.status_code == 200:
        return {"ok": True, "model": model}
    raise HTTPException(status_code=400, detail=f"Key test failed ({resp.status_code}): {resp.text[:200]}")

# ── Import ────────────────────────────────────────────────────
from fastapi import UploadFile, File
from fastapi.responses import StreamingResponse
import io, openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

@app.get("/import/template")
def download_import_template(user=Depends(current_user)):
    """Generate and return a filled-out XLSX import template."""
    wb = openpyxl.Workbook()

    # ── Log sheet ──
    ws = wb.active
    ws.title = "Log"

    headers = ["Date", "Bodyweight (lbs)", "Exercise", "Tool", "Modifier", "Load", "Reps", "Time (min)", "Phase"]
    header_fill = PatternFill("solid", fgColor="1A1A18")
    header_font = Font(bold=True, color="E05A20")
    thin = Side(style="thin", color="444444")
    border = Border(bottom=Side(style="thin", color="444444"))

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center")
        cell.border = border

    # Column widths
    widths = [14, 18, 28, 14, 20, 10, 8, 12, 16]
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    # Example rows
    examples = [
        ["01/15/2026", 200.0, "Bench Press",     "Bar",        "Standard",  185, 5,  1.5],
        ["01/15/2026", 200.0, "Bench Press",     "Bar",        "Standard",  185, 5,  1.5],
        ["01/15/2026", 200.0, "Bench Press",     "Bar",        "Standard",  185, 5,  1.5],
        ["01/15/2026", 200.0, "Incline Fly",     "Cable",      "Double",    40,  12, 1.5],
        ["01/15/2026", 200.0, "Incline Fly",     "Cable",      "Double",    40,  12, 1.5],
        ["01/15/2026", 200.0, "Pull-up",         "Bodyweight", "Vest",      25,  8,  1.5],
        ["01/15/2026", 200.0, "Pull-up",         "Bodyweight", "Vest",      25,  8,  1.5],
        ["01/17/2026", 199.5, "Low Bar Squat",   "Bar",        "Standard",  225, 3,  2.0],
        ["01/17/2026", 199.5, "KB Swings",       "Kettlebell", "",          53,  15, 1.5, "Recomp"],
    ]
    example_font = Font(color="888888", italic=True)
    for r, row in enumerate(examples, 2):
        for c, val in enumerate(row, 1):
            cell = ws.cell(row=r, column=c, value=val)
            cell.font = example_font

    ws.freeze_panes = "A2"

    # ── Legend sheet ──
    wl = wb.create_sheet("Legend")
    wl.column_dimensions["A"].width = 20
    wl.column_dimensions["B"].width = 60
    wl.column_dimensions["C"].width = 50

    legend_title_font = Font(bold=True, color="E05A20", size=13)
    section_font = Font(bold=True, color="CCCCCC")
    body_font = Font(color="AAAAAA")
    wl["A1"] = "Agon Import — Field Legend"
    wl["A1"].font = legend_title_font

    rows = [
        ("", "", ""),
        ("FIELD", "DESCRIPTION", "VALID VALUES / NOTES"),
        ("Date", "Session date", "MM/DD/YYYY — all sets for a session share the same date"),
        ("Bodyweight (lbs)", "Your bodyweight that day", "Decimal OK — e.g. 199.5"),
        ("Exercise", "Exercise name", "Any name. New exercises will be created automatically."),
        ("Tool", "Equipment type", "Bar | Cable | Kettlebell | Dumbbell | Machine | Bodyweight"),
        ("Modifier", "Sub-type of tool or bodyweight mode", "See below"),
        ("Load", "Weight entered (see modifier rules)", "Numeric. See below for what to enter per modifier."),
        ("Reps", "Reps performed in this set", "Integer"),
        ("Time (min)", "Duration of this set in minutes", "Decimal. e.g. 1.5 for 90 seconds. Defaults to 1.5 if left blank."),
        ("Phase", "Training phase active on this date (optional)", "Weight Loss | Cut | Recomp | Bulk | Maintain | Peak | Deload. Leave blank to skip. Consecutive rows with the same phase are grouped into one phase record."),
        ("", "", ""),
        ("MODIFIER GUIDE", "", ""),
        ("Tool: Bar", "", ""),
        ("  Standard", "Two-sided barbell, enter total weight", "Load = total bar weight (e.g. 135 for 45lb bar + 45lb per side)"),
        ("  Vitruvian", "Vitruvian bar, enter weight per side", "Load = plates per side (trueLbs = Load × 2)"),
        ("Tool: Cable", "", ""),
        ("  Single", "One cable / one arm", "Load = stack weight (trueLbs = Load × 1)"),
        ("  Double", "Two cables / both arms", "Load = stack weight per side (trueLbs = Load × 2)"),
        ("Tool: Dumbbell", "", ""),
        ("  (leave blank)", "Always counts both dumbbells", "Load = weight of one dumbbell (trueLbs = Load × 2)"),
        ("Tool: Kettlebell", "", ""),
        ("  (leave blank)", "Single bell", "Load = bell weight (trueLbs = Load × 1)"),
        ("Tool: Machine", "", ""),
        ("  (leave blank)", "Machine stack", "Load = stack weight (trueLbs = Load × 1)"),
        ("Tool: Bodyweight", "", ""),
        ("  Normal", "Just bodyweight", "Load = 0 or leave blank (trueLbs = bodyweight)"),
        ("  Vest", "Bodyweight + weighted vest", "Load = vest weight in lbs (trueLbs = bodyweight + vest)"),
        ("  Assisted", "Bodyweight with band/machine assist", "Load = assist % (e.g. 20 = 20% assist, trueLbs = bodyweight × 0.80)"),
        ("", "", ""),
        ("NOTES", "", ""),
        ("", "Each ROW is one SET.", "Group sets for the same exercise on the same date together."),
        ("", "Time per set defaults to 1.5 min.", "All imported sets use 1.5 min for RD calculation."),
        ("", "Existing sessions on the same date will NOT be overwritten.", "Change the date or delete the existing session first."),
        ("", "Grey rows in the Log sheet are examples — delete them before importing.", ""),
    ]

    for r, (a, b, c) in enumerate(rows, 2):
        wl.cell(row=r, column=1, value=a)
        wl.cell(row=r, column=2, value=b)
        wl.cell(row=r, column=3, value=c)
        if a in ("FIELD", "MODIFIER GUIDE", "NOTES"):
            for col in range(1, 4):
                wl.cell(row=r, column=col).font = section_font
        elif a in ("Tool: Bar", "Tool: Cable", "Tool: Dumbbell", "Tool: Kettlebell", "Tool: Machine", "Tool: Bodyweight"):
            for col in range(1, 4):
                wl.cell(row=r, column=col).font = Font(bold=True, color="E05A20")
        else:
            for col in range(1, 4):
                wl.cell(row=r, column=col).font = body_font

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return StreamingResponse(
        buf,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=agon_import_template.xlsx"}
    )


def _parse_import_row(row, bw: float) -> dict:
    """Convert a single spreadsheet row into a set dict with optional phase."""
    tool      = (str(row[3] or "Bar")).strip()
    modifier  = (str(row[4] or "")).strip().lower()
    raw_load  = float(row[5] or 0)
    reps      = int(float(row[6] or 0))
    time_min  = float(row[7]) if len(row) > 7 and row[7] not in (None, "", " ") else 1.5
    phase_raw = row[8] if len(row) > 8 else None
    phase     = str(phase_raw).strip() if phase_raw not in (None, "", " ", "None") else None

    # Determine trueLbs based on tool + modifier
    if tool == "Bar":
        if "vitruvian" in modifier:
            true_lbs = raw_load * 2
            mult = 2
        else:  # Standard
            true_lbs = raw_load
            mult = 2
    elif tool == "Cable":
        if "double" in modifier:
            true_lbs = raw_load * 2
            mult = 2
        else:  # Single
            true_lbs = raw_load
            mult = 1
    elif tool == "Dumbbell":
        true_lbs = raw_load * 2
        mult = 2
    elif tool == "Kettlebell":
        true_lbs = raw_load
        mult = 1
    elif tool == "Machine":
        true_lbs = raw_load
        mult = 1
    elif tool == "Bodyweight":
        if "vest" in modifier:
            true_lbs = bw + raw_load
        elif "assisted" in modifier or "assist" in modifier:
            true_lbs = bw * (1 - raw_load / 100.0)
        else:
            true_lbs = bw
        mult = 1
    else:
        true_lbs = raw_load
        mult = 1

    vol = round(true_lbs * reps, 2)
    return {
        "rawLoad": raw_load,
        "trueLbs": round(true_lbs, 2),
        "reps": reps,
        "vol": vol,
        "time":  time_min,
        "phase": phase
    }


@app.post("/import/sessions")
async def import_sessions(user=Depends(current_user), file: UploadFile = File(...)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot import data")

    contents = await file.read()
    try:
        wb = openpyxl.load_workbook(io.BytesIO(contents), data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(status_code=400, detail=f"Could not read XLSX: {str(e)[:200]}")

    # Parse rows (skip header row 1)
    from collections import defaultdict, OrderedDict
    sessions_map = OrderedDict()  # key: (date, bw) → {exercises: {name → {tool, sets[]}}}

    skipped = []
    errors  = []

    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        if not any(row):
            continue
        try:
            raw_date = str(row[0] or "").strip()
            bw       = float(row[1] or 0)
            ex_name  = str(row[2] or "").strip()
            tool     = str(row[3] or "Bar").strip()
            modifier = str(row[4] or "").strip()

            if not raw_date or not ex_name or not bw:
                continue

            # Normalise date to YYYY-MM-DD
            from datetime import datetime
            for fmt in ("%m/%d/%Y", "%Y-%m-%d", "%m-%d-%Y", "%d/%m/%Y"):
                try:
                    date_str = datetime.strptime(raw_date, fmt).strftime("%Y-%m-%d")
                    break
                except ValueError:
                    continue
            else:
                errors.append(f"Row {row_idx}: unrecognised date format '{raw_date}'")
                continue

            set_data = _parse_import_row(row, bw)
            if set_data["reps"] <= 0:
                continue

            key = (date_str, bw)
            if key not in sessions_map:
                sessions_map[key] = {}
            if ex_name not in sessions_map[key]:
                sessions_map[key][ex_name] = {"tool": tool, "modifier": modifier, "sets": []}
            sessions_map[key][ex_name]["sets"].append(set_data)

        except Exception as e:
            errors.append(f"Row {row_idx}: {str(e)[:100]}")

    if not sessions_map:
        raise HTTPException(status_code=400, detail="No valid rows found. Check the template format.")

    # Build session objects and insert
    conn = get_db()
    inserted = 0
    skipped_dates = []

    for (date_str, bw), exercises_map in sessions_map.items():
        # Check for existing session
        existing = conn.execute(
            "SELECT id FROM sessions WHERE user_id=? AND date=?",
            (user["id"], date_str)
        ).fetchone()
        if existing:
            skipped_dates.append(date_str)
            continue

        exercises = []
        for ex_name, ex_data in exercises_map.items():
            sets = ex_data["sets"]
            total_vol  = sum(s["vol"]  for s in sets)
            total_time = sum(s["time"] for s in sets)
            density = round(total_vol / total_time, 2) if total_time > 0 else 0
            exercises.append({
                "name":     ex_name,
                "tool":     ex_data["tool"],
                "sets":     sets,
                "totalVol": round(total_vol, 2),
                "density":  density,
                "isBW":     ex_data["tool"] == "Bodyweight"
            })

        total_vol  = sum(s["vol"]  for ex in exercises for s in ex["sets"])
        total_time = sum(s["time"] for ex in exercises for s in ex["sets"])
        rd = round(total_vol / (total_time * bw), 4) if (total_time > 0 and bw) else 0
        total_density = round(total_vol / total_time, 4) if total_time > 0 else 0

        conn.execute(
            "INSERT INTO sessions (user_id, date, bw, rd, total_density, exercises) VALUES (?,?,?,?,?,?)",
            (user["id"], date_str, bw, rd, total_density, json.dumps(exercises))
        )
        inserted += 1

    # Synthesize phases from Phase column — group consecutive same-phase dates
    date_phase_map = {}  # date -> phase_type
    for (d_str, bw_val), ex_map in sessions_map.items():
        for ex_name, ex_data in ex_map.items():
            for s in ex_data["sets"]:
                if s.get("phase"):
                    date_phase_map[d_str] = s["phase"]
                    break
            break

    phases_created = 0
    if date_phase_map:
        sorted_dates = sorted(date_phase_map.keys())
        runs = []  # [(phase_type, start, end)]
        cur_type  = date_phase_map[sorted_dates[0]]
        cur_start = sorted_dates[0]
        prev_date = sorted_dates[0]
        for d in sorted_dates[1:]:
            ptype = date_phase_map[d]
            if ptype != cur_type:
                runs.append((cur_type, cur_start, prev_date))
                cur_type  = ptype
                cur_start = d
            prev_date = d
        runs.append((cur_type, cur_start, None))  # last run stays open

        for phase_type, start_date, end_date in runs:
            if not phase_type:
                continue
            existing = conn.execute(
                "SELECT id FROM phases WHERE user_id=? AND phase_type=? AND start_date=?",
                (user["id"], phase_type, start_date)
            ).fetchone()
            if not existing:
                conn.execute(
                    "INSERT INTO phases (user_id, phase_type, start_date, end_date) VALUES (?,?,?,?)",
                    (user["id"], phase_type, start_date, end_date)
                )
                phases_created += 1

    conn.commit()
    conn.close()

    return {
        "inserted":       inserted,
        "skipped":        skipped_dates,
        "errors":         errors,
        "phases_created": phases_created,
    }

# ── Phase endpoints ───────────────────────────────────────────
@app.get("/phases")
def get_phases(user=Depends(current_user)):
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT id, phase_type, start_date, end_date, notes FROM phases WHERE user_id=? ORDER BY start_date DESC",
        (uid,)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/phases")
def create_phase(body: PhaseIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot create phases")
    from datetime import date as _date
    today_str = _date.today().strftime("%Y-%m-%d")
    conn = get_db()
    # Only auto-close the open phase if new phase starts today or in future
    # For historical phases (past start date) the user sets end dates manually
    if not body.end_date and body.start_date >= today_str:
        conn.execute(
            "UPDATE phases SET end_date=? WHERE user_id=? AND end_date IS NULL",
            (body.start_date, user["id"])
        )
    conn.execute(
        "INSERT INTO phases (user_id, phase_type, start_date, end_date, notes) VALUES (?,?,?,?,?)",
        (user["id"], body.phase_type, body.start_date, body.end_date or None, body.notes or None)
    )
    conn.commit()
    conn.close()
    return {"status": "created"}

@app.delete("/phases/{phase_id}")
def delete_phase(phase_id: int, user=Depends(current_user)):
    conn = get_db()
    conn.execute("DELETE FROM phases WHERE id=? AND user_id=?", (phase_id, user["id"]))
    conn.commit()
    conn.close()
    return {"status": "deleted"}

@app.put("/phases/{phase_id}/close")
def close_phase(phase_id: int, user=Depends(current_user)):
    from datetime import date as _date
    today = _date.today().strftime("%Y-%m-%d")
    conn = get_db()
    conn.execute(
        "UPDATE phases SET end_date=? WHERE id=? AND user_id=? AND end_date IS NULL",
        (today, phase_id, user["id"])
    )
    conn.commit()
    conn.close()
    return {"status": "closed"}


# ── Recomposition score ───────────────────────────────────────
@app.get("/recomp")
def get_recomp(user=Depends(current_user)):
    """
    Recomp score: measures whether RD is rising while BW is falling.
    Uses last 8 weeks of data. Returns a score and trend interpretation.
    Score: positive = gaining strength while losing weight (ideal recomp)
           near zero = maintaining
           negative  = losing strength relative to bodyweight
    """
    uid = data_user_id(user)
    conn = get_db()
    rows = conn.execute(
        "SELECT date, bw, rd FROM sessions WHERE user_id=? ORDER BY date DESC LIMIT 60",
        (uid,)
    ).fetchall()
    # Get target_bw — defensive in case column not yet migrated
    try:
        settings = conn.execute("SELECT target_bw FROM user_settings WHERE user_id=?", (uid,)).fetchone()
    except Exception:
        settings = None
    conn.close()

    if len(rows) < 4:
        return {"score": None, "label": "Not enough data", "bw_trend": None, "rd_trend": None, "target_bw": settings["target_bw"] if settings else None}

    rows = list(reversed(rows))  # chronological
    # Split into two halves
    mid = len(rows) // 2
    first_half  = rows[:mid]
    second_half = rows[mid:]

    avg_rd_early = sum(r["rd"] for r in first_half)  / len(first_half)
    avg_rd_late  = sum(r["rd"] for r in second_half) / len(second_half)
    avg_bw_early = sum(r["bw"] for r in first_half)  / len(first_half)
    avg_bw_late  = sum(r["bw"] for r in second_half) / len(second_half)

    bw_trend = round(avg_bw_late - avg_bw_early, 1)   # negative = losing weight
    rd_trend = round(avg_rd_late - avg_rd_early, 3)    # positive = getting stronger

    # Score: RD change normalised, penalised if BW is rising (bulk is fine but label differently)
    score = round(rd_trend * 100, 1)

    if rd_trend > 0.02 and bw_trend < 0:
        label = "Recomping"
    elif rd_trend > 0.02 and bw_trend >= 0:
        label = "Building"
    elif abs(rd_trend) <= 0.02 and bw_trend < -0.5:
        label = "Cutting"
    elif abs(rd_trend) <= 0.02:
        label = "Maintaining"
    else:
        label = "Regressing"

    # Current bw vs target
    current_bw = rows[-1]["bw"] if rows else None
    target_bw  = settings["target_bw"] if settings else None
    to_goal    = round(current_bw - target_bw, 1) if (current_bw and target_bw) else None

    return {
        "score":      score,
        "label":      label,
        "bw_trend":   bw_trend,
        "rd_trend":   rd_trend,
        "current_bw": current_bw,
        "target_bw":  target_bw,
        "to_goal":    to_goal,
    }

# ── Protocol endpoints ────────────────────────────────────────
@app.get("/protocols")
def get_protocols(user=Depends(current_user)):
    conn = get_db()
    rows = conn.execute(
        "SELECT id, name, dose, frequency, notes FROM protocols WHERE user_id=? ORDER BY sort_order, id",
        (user["id"],)
    ).fetchall()
    conn.close()
    return [dict(r) for r in rows]

@app.post("/protocols")
def add_protocol(body: ProtocolIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot add protocols")
    conn = get_db()
    max_order = conn.execute(
        "SELECT COALESCE(MAX(sort_order),0) FROM protocols WHERE user_id=?", (user["id"],)
    ).fetchone()[0]
    conn.execute(
        "INSERT INTO protocols (user_id, name, dose, frequency, notes, sort_order) VALUES (?,?,?,?,?,?)",
        (user["id"], body.name.strip(), body.dose, body.frequency, body.notes, max_order + 1)
    )
    conn.commit()
    conn.close()
    return {"status": "added"}

@app.put("/protocols/{protocol_id}")
def update_protocol(protocol_id: int, body: ProtocolIn, user=Depends(current_user)):
    if user["role"] == "demo":
        raise HTTPException(status_code=403, detail="Demo accounts cannot modify protocols")
    conn = get_db()
    conn.execute(
        "UPDATE protocols SET name=?, dose=?, frequency=?, notes=? WHERE id=? AND user_id=?",
        (body.name.strip(), body.dose, body.frequency, body.notes, protocol_id, user["id"])
    )
    conn.commit()
    conn.close()
    return {"status": "updated"}

@app.delete("/protocols/{protocol_id}")
def delete_protocol(protocol_id: int, user=Depends(current_user)):
    conn = get_db()
    conn.execute("DELETE FROM protocols WHERE id=? AND user_id=?", (protocol_id, user["id"]))
    conn.commit()
    conn.close()
    return {"status": "deleted"}
